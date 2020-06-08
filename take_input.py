from __future__ import print_function
import openpyxl as xl
import time


def Read_Onboard_Sheet(File_Name,index):
    wb= xl.load_workbook(File_Name)
    sheet = wb["ISE"]
    return sheet

def numberofvm(sheet):

    rows = sheet.max_row
    vmcount = rows-1

    return vmcount

def gethostname(sheet,vmcount):

    rows = vmcount+2
    column = 1
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def getesxiname(sheet,vmcount):

    rows = vmcount + 2
    column = 3
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def getdsname(sheet,vmcount):

    rows = vmcount + 2
    column = 5
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def getdcname(sheet, vmcount):

    rows = vmcount + 2
    column = 6
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def getcpu(sheet, vmcount):

    cpu = [6,8,8,8,12,12]
    rows = vmcount + 2
    column = 2
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list

        if sheet.cell(row=i, column=column).value == 3515:
            static_variable.append(cpu[0])
        if sheet.cell(row=i, column=column).value == 3555:
            static_variable.append(cpu[1])
        if sheet.cell(row=i, column=column).value == 3595:
            static_variable.append(cpu[2])
        if sheet.cell(row=i, column=column).value == 3615:
            static_variable.append(cpu[3])
        if sheet.cell(row=i, column=column).value == 3655:
            static_variable.append(cpu[4])
        if sheet.cell(row=i, column=column).value == 3695:
            static_variable.append(cpu[5])

    return static_variable

def getmemory(sheet, vmcount):

    memory = [17000, 33000, 65000, 33000, 97000, 257000]
    rows = vmcount + 2
    column = 2
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list

        if sheet.cell(row=i, column=column).value == 3515:
            static_variable.append(memory[0])
        if sheet.cell(row=i, column=column).value == 3555:
            static_variable.append(memory[1])
        if sheet.cell(row=i, column=column).value == 3595:
            static_variable.append(memory[2])
        if sheet.cell(row=i, column=column).value == 3615:
            static_variable.append(memory[3])
        if sheet.cell(row=i, column=column).value == 3655:
            static_variable.append(memory[4])
        if sheet.cell(row=i, column=column).value == 3695:
            static_variable.append(memory[5])

    return static_variable

def getdisk(vmcount):
    disk = [200, 200, 2400, 2400, 600, 600]
    rows = vmcount + 2
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(disk[i-2])
    return static_variable

def getguestver(sheet, vmcount):

    memory = ["vmx-13", "vmx-14", "vmx-15", "vmx-17"]
    rows = vmcount + 2
    column = 4
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list

        if sheet.cell(row=i, column=column).value == 6.5:
            static_variable.append(memory[0])
        if sheet.cell(row=i, column=column).value == 6.7:
            static_variable.append(memory[1])
        if sheet.cell(row=i, column=column).value == 6.8:
            static_variable.append(memory[2])
        if sheet.cell(row=i, column=column).value == 7.0:
            static_variable.append(memory[3])

    return static_variable

def getisods(sheet):

    column = 7
    row = 2
    static_variable = sheet.cell(row=row, column=column).value

    return static_variable

def getisofile(sheet):
    column = 8
    row = 2
    static_variable = sheet.cell(row=row, column=column).value

    return static_variable

def getnetwork1(sheet,vmcount):
    rows = vmcount + 2
    column = 9
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def getnetwork2(sheet,vmcount):
    rows = vmcount + 2
    column = 10
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable


def getmgmtip(sheet,vmcount):
    rows = vmcount + 2
    column = 11
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def getmgmtmask(sheet,vmcount):
    rows = vmcount + 2
    column = 12
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def getmgmtgateway(sheet,vmcount):
    rows = vmcount + 2
    column = 13
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def getdomain(sheet,vmcount):
    rows = vmcount + 2
    column = 14
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def getnameserver1(sheet,vmcount):
    rows = vmcount + 2
    column = 15
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def getnameserver2(sheet,vmcount):
    rows = vmcount + 2
    column = 16
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def getntpserver1(sheet,vmcount):
    rows = vmcount + 2
    column = 17
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def getntpserver2(sheet,vmcount):
    rows = vmcount + 2
    column = 18
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable

def gettimezone(sheet,vmcount):
    rows = vmcount + 2
    column = 19
    static_variable = []
    for i in range(2, rows):
        # reading cell value from source excel file and put it in a list
        static_variable.append(sheet.cell(row=i, column=column).value)
    return static_variable


def countdown(t):
    while t:
        mins, secs = divmod(t, 60)
        timeformat = '{:02d}:{:02d}'.format(mins, secs)
        print (timeformat, end='\r')
        time.sleep(1)
        t -= 1
        if t == 10800:
            print("ISE Installation is still going on : 3 hours to go")
        if t == 7200:
            print("ISE Installation is still going on : 2 hours to go")
        if t == 3600:
            print("ISE Installation is still going on : 1 hour to go")
        if t == 1800:
            print("ISE Installation is still going on : 30 minutes to go")
        if t == 900:
            print("ISE Installation is still going on : 15 minutes to go")
        if t == 300:
            print("ISE Installation is still going on : 5 minutes to go")


    print('ISE Installation Complete - Sending Day-0 Configuration\n')


'''
Dict for Character to Keystroke 
'''
keycodeDict = {"a": "KEY_A", "b": "KEY_B", "c": "KEY_C", "d": "KEY_D", "e": "KEY_E", "f": "KEY_F", "g": "KEY_G",
               "h": "KEY_H", "i": "KEY_I", "j": "KEY_J", "k": "KEY_K", "l": "KEY_L", "m": "KEY_M", "n": "KEY_N",
               "o": "KEY_O", "p": "KEY_P", "q": "KEY_Q", "r": "KEY_R", "s": "KEY_S", "t": "KEY_T", "u": "KEY_U",
               "v": "KEY_V", "w": "KEY_W", "x": "KEY_X", "y": "KEY_Y", "z": "KEY_Z", "1": "KEY_1", "2": "KEY_2",
               "3": "KEY_3", "4": "KEY_4", "5": "KEY_5", "6": "KEY_6", "7": "KEY_7", "8": "KEY_8",
               "9": "KEY_9", "0": "KEY_0", "-": "KEY_MINUS", "{": "KEY_LEFTBRACE", ".": "KEY_DOT",
               "A": ("KEY_CAPSLOCK", 'KEY_A', "KEY_CAPSLOCK"), "B": ("KEY_CAPSLOCK", 'KEY_B', "KEY_CAPSLOCK"),
               "C": ("KEY_CAPSLOCK", 'KEY_C', "KEY_CAPSLOCK"), "/": "KEY_SLASH"
               }


def strtokey(key):
    h =[]
    vars = list(key)
    for var in vars:
        h.append(keycodeDict[var])
    h.append('KEY_ENTER')
    return h
